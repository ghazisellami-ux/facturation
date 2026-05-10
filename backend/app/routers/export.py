"""
Router pour l'export Excel/CSV des factures et retenues.
"""
import io
import csv
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.models.user import User
from app.models.company import Company
from app.models.client import Client
from app.models.supplier import Supplier
from app.models.invoice import Invoice, InvoiceType
from app.models.withholding import WithholdingTax
from app.utils.auth import get_current_user_from_token
from app.routers.invoices import get_user_company

router = APIRouter(prefix="/api/export", tags=["Export"])

MONTHS_FR = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
]

STATUS_LABELS = {
    "brouillon": "Brouillon",
    "envoyee": "Envoyée",
    "payee": "Payée",
    "partiellement_payee": "Partielle",
    "en_retard": "En retard",
    "annulee": "Annulée",
}


def _get_invoices(db: Session, company_id: str, inv_type: str, year: int, month: Optional[int]):
    """Fetch invoices filtered by type, year, and optionally month."""
    q = db.query(Invoice).filter(
        Invoice.company_id == company_id,
        Invoice.invoice_type == inv_type,
        extract("year", Invoice.date) == year,
    )
    if month:
        q = q.filter(extract("month", Invoice.date) == month)
    return q.order_by(Invoice.date.asc()).all()


def _get_withholdings(db: Session, company_id: str, year: int, month: Optional[int]):
    """Fetch withholdings filtered by year, and optionally month."""
    q = db.query(WithholdingTax).filter(
        WithholdingTax.company_id == company_id,
        extract("year", WithholdingTax.date) == year,
    )
    if month:
        q = q.filter(extract("month", WithholdingTax.date) == month)
    return q.order_by(WithholdingTax.date.asc()).all()


def _resolve_name(db: Session, inv: Invoice) -> str:
    if inv.client_id:
        c = db.query(Client).filter(Client.id == inv.client_id).first()
        return c.name if c else ""
    if inv.supplier_id:
        s = db.query(Supplier).filter(Supplier.id == inv.supplier_id).first()
        return s.name if s else ""
    return ""


def _resolve_wh_name(db: Session, w: WithholdingTax) -> str:
    if w.client_id:
        c = db.query(Client).filter(Client.id == w.client_id).first()
        return c.name if c else w.beneficiary_name or ""
    if w.supplier_id:
        s = db.query(Supplier).filter(Supplier.id == w.supplier_id).first()
        return s.name if s else w.beneficiary_name or ""
    return w.beneficiary_name or ""


def _period_label(year: int, month: Optional[int]) -> str:
    if month:
        return f"{MONTHS_FR[month - 1]} {year}"
    return str(year)


# ─── FACTURES VENTE ───

@router.get("/factures")
def export_factures(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    format: str = Query("xlsx"),
    token: str = Query(...),
    db: Session = Depends(get_db),
):
    current_user = get_current_user_from_token(token, db)
    company = get_user_company(db, current_user)
    invoices = _get_invoices(db, company.id, InvoiceType.FACTURE.value, year, month)
    period = _period_label(year, month)

    headers = ["Réf.", "Date", "Client", "Total HT", "TVA", "Total TTC", "Payé", "Reste", "Statut"]

    rows = []
    for inv in invoices:
        name = _resolve_name(db, inv)
        rows.append([
            inv.reference,
            inv.date.strftime("%d/%m/%Y") if inv.date else "",
            name,
            round(inv.subtotal or 0, 3),
            round(inv.tva_amount or 0, 3),
            round(inv.total or 0, 3),
            round(inv.amount_paid or 0, 3),
            round(inv.balance_due or 0, 3),
            STATUS_LABELS.get(inv.status, inv.status),
        ])

    title = f"Factures de vente — {period}"
    filename = f"factures_vente_{year}_{month or 'annuel'}"

    if format == "csv":
        return _csv_response(headers, rows, filename)
    return _xlsx_response(headers, rows, title, filename)


# ─── FACTURES ACHAT ───

@router.get("/achats")
def export_achats(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    format: str = Query("xlsx"),
    token: str = Query(...),
    db: Session = Depends(get_db),
):
    current_user = get_current_user_from_token(token, db)
    company = get_user_company(db, current_user)
    invoices = _get_invoices(db, company.id, InvoiceType.FACTURE_ACHAT.value, year, month)
    period = _period_label(year, month)

    headers = ["Réf.", "Date", "Fournisseur", "Total HT", "TVA", "Total TTC", "Payé", "Reste", "Statut"]

    rows = []
    for inv in invoices:
        name = _resolve_name(db, inv)
        rows.append([
            inv.reference,
            inv.date.strftime("%d/%m/%Y") if inv.date else "",
            name,
            round(inv.subtotal or 0, 3),
            round(inv.tva_amount or 0, 3),
            round(inv.total or 0, 3),
            round(inv.amount_paid or 0, 3),
            round(inv.balance_due or 0, 3),
            STATUS_LABELS.get(inv.status, inv.status),
        ])

    title = f"Factures d'achat — {period}"
    filename = f"factures_achat_{year}_{month or 'annuel'}"

    if format == "csv":
        return _csv_response(headers, rows, filename)
    return _xlsx_response(headers, rows, title, filename)


# ─── RETENUES ───

@router.get("/retenues")
def export_retenues(
    year: int = Query(...),
    month: Optional[int] = Query(None),
    format: str = Query("xlsx"),
    token: str = Query(...),
    db: Session = Depends(get_db),
):
    current_user = get_current_user_from_token(token, db)
    company = get_user_company(db, current_user)
    items = _get_withholdings(db, company.id, year, month)
    period = _period_label(year, month)

    headers = ["Date", "Référence", "Type", "Bénéficiaire", "MF", "Taux", "Base HT", "Montant retenu"]

    rows = []
    for w in items:
        name = _resolve_wh_name(db, w)
        rows.append([
            w.date.strftime("%d/%m/%Y") if w.date else "",
            w.reference or "",
            "Émise" if w.type == "emise" else "Reçue",
            name,
            w.beneficiary_tax_id or "",
            f"{w.rate}%",
            round(w.base_amount or 0, 3),
            round(w.tax_amount or 0, 3),
        ])

    title = f"Retenues à la source — {period}"
    filename = f"retenues_{year}_{month or 'annuel'}"

    if format == "csv":
        return _csv_response(headers, rows, filename)
    return _xlsx_response(headers, rows, title, filename)


# ─── Helpers ───

def _csv_response(headers: list, rows: list, filename: str) -> StreamingResponse:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def _xlsx_response(headers: list, rows: list, title: str, filename: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1565C0")
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    number_fmt = '#,##0.000'

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    # Headers at row 3
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data
    for r_idx, row in enumerate(rows, 4):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = number_fmt
                cell.alignment = Alignment(horizontal="right")

    # Auto-width
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)

    # Totals row
    if rows:
        total_row = len(rows) + 4
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
        for c_idx, val in enumerate(rows[0], 1):
            if isinstance(val, (int, float)):
                total = sum(r[c_idx - 1] for r in rows if isinstance(r[c_idx - 1], (int, float)))
                cell = ws.cell(row=total_row, column=c_idx, value=round(total, 3))
                cell.font = Font(bold=True, size=11)
                cell.number_format = number_fmt
                cell.border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
